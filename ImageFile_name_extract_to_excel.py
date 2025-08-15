import os
from openpyxl import Workbook

# Path to your folder containing images
folder_path = r"E:\Images"

# Create a new Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Image File Names"

# Add headers
ws['A1'] = "File Name"
ws['B1'] = "Full Path"

# List all files in folder
row = 2
for file_name in os.listdir(folder_path):
    if file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff')):
        ws.cell(row=row, column=1).value = file_name
        ws.cell(row=row, column=2).value = os.path.join(folder_path, file_name)
        row += 1

# Save the Excel file
output_file = r"E:\Images"
wb.save(output_file)

print(f"Excel file created: {output_file}")
